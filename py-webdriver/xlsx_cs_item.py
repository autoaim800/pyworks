import os
import sys
import sqlite3
import openpyxl
from openpyxl.styles import Style, Color, Font

DDL = """
    engine varchar(40),
    body varchar(10),
    title varchar(120),
        recordid varchar(40),
        price float,
        odometer int,
    state varchar(3),
    href text,
    trans varchar(20),
        year varchar(4),
        brand varchar(20),
        model varchar(20),
    badge varchar(20),
    variant varchar(40),
    year_cost int,
        kkms_cost int,
        seller varchar(20),
        cr_date date,    
    drive_type varchar(40),      
    pri_img text,    
    series varchar(40)
    
    primary key (recordid)
"""

HEADERS_ROW = [("recordid", 3),
           ("year", 9), ("brand", 10), ("model", 11), \
           ("badge", 12), ("price", 4), ("odometer", 5), \
           ("title", 2), ("variant", 13), ("state", 6), \
           ("trans", 8), ("engine", 0), ("body", 1), \
           ("href", 7), ("year-cost", 14), ("kkms-cost", 15), \
           ("seller", 16), ("cr_date", 17), ("drive_type", 18), \
           ("series", 20), ("offset_price", 21), ("pri_img", 19)]

HEADERS_XL = ["cr_date", "href", "year", "brand", "model", "badge", "odometer", \
          "price", "year-cost", "kkms-cost", "offset_price", "title", "series", \
          "variant", "seller", "state", \
          "trans", "engine", "body", "drive_type", "recordid"]

QSQL_CMD = """
select *
from t_item
where price > 0 and kkms_cost > 0 and year >=2003
-- and brand = 'Holden' and model = 'Commodore'
order by offset_price
-- ,price, kkms_cost, year_cost
"""

class XlsxCs:

    def __init__(self, dbFp):
        self.db = sqlite3.connect(dbFp)

    def closeDb(self):
        self.db.close()

    def saveAs(self, outFp):
        wb = openpyxl.workbook.Workbook()
        ws = wb.get_active_sheet()
        rowIndex = 0
        # write header
        colIndex = 0
        for head in HEADERS_XL:
            cell = ws.cell(row=rowIndex + 1, column=colIndex + 1)
            cell.value = head
            cell.style = Style(font=Font(bold=True))
            colIndex += 1
        # write content
        for row in self.db.execute(QSQL_CMD):
            rowIndex += 1
            for head, index in HEADERS_ROW:
                colIndex = HEADERS_XL.index(head)
                if "href" == head:
                    cell = ws.cell(row=rowIndex + 1, column=colIndex + 1)
                    cell.value = row[index]
                    cell.hyperlink = row[index]
                else:
                    ws.cell(row=rowIndex + 1, column=colIndex + 1).value = row[index]
            if rowIndex % 100 == 0:
                sys.stdout.write('.')

        ws.auto_filter.ref = "A1:S1"
        ws.freeze_panes = ws.cell('A2')

        print("writing...")
        wb.save(outFp)

def main():
    dbFp = sys.argv[1]
    outFp = sys.argv[2]

    xc = XlsxCs(dbFp)
    xc.saveAs(outFp)
    xc.closeDb()

if __name__ == '__main__':
    exit(main())
